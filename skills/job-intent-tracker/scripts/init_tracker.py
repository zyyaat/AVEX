#!/usr/bin/env python3
"""
init_tracker.py — 初始化求职投递追踪表

用法：
    python init_tracker.py --format xlsx --output tracker.xlsx
    python init_tracker.py --format md   --output tracker.md

xlsx 模式生成带表头、下拉验证、条件格式的 Excel。
md 模式生成轻量的 Markdown 表格，便于在 Notion / Obsidian / 飞书文档里粘贴。
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

COLUMNS = [
    "公司",
    "岗位",
    "来源",          # 猎头 / 官网 / 内推 / 招聘网站 / 其他
    "JD链接",
    "投递日期",
    "当前阶段",       # 投递 / 笔试 / 一面 / 二面 / 三面 / HR面 / Offer / Reject / 沉默
    "下一步动作",
    "Deadline",
    "薪资范围",
    "内推人",
    "备注",
]

STAGE_OPTIONS = [
    "投递", "笔试", "一面", "二面", "三面", "HR面", "Offer", "Reject", "沉默",
]

SOURCE_OPTIONS = ["官网", "内推", "招聘网站", "猎头", "校招", "其他"]


def write_md(path: Path) -> None:
    header = "| " + " | ".join(COLUMNS) + " |"
    sep = "| " + " | ".join(["---"] * len(COLUMNS)) + " |"
    sample = "| 示例公司 | 高级产品经理 | 内推 | https://... | 2026-05-07 | 投递 | 等HR反馈 | 2026-05-14 | 40-55k | 张三 | 内推码 ABC123 |"
    legend_lines = [
        "",
        "## 字段说明",
        "",
        f"- **来源**：{' / '.join(SOURCE_OPTIONS)}",
        f"- **当前阶段**：{' / '.join(STAGE_OPTIONS)}",
        "- **下一步动作**：写一句具体可执行的事，比如『5/10 前发感谢信』『等 HR 回复，5/15 未回则催』",
        "- **沉默**：超过 7 天没反馈的状态，提醒自己主动 follow up 或放弃",
        "",
    ]
    content = "\n".join([
        "# 投递追踪表",
        "",
        header,
        sep,
        sample,
        *legend_lines,
    ])
    path.write_text(content, encoding="utf-8")
    print(f"✓ Markdown 追踪表已生成：{path}")


def write_xlsx(path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import CellIsRule
    except ImportError:
        print(
            "✗ 缺少 openpyxl，请先安装：pip install openpyxl --break-system-packages",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "投递追踪"

    # 表头
    ws.append(COLUMNS)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    widths = [16, 22, 10, 28, 12, 10, 26, 12, 14, 10, 30]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w

    # 示例行
    sample_row = [
        "示例公司",
        "高级产品经理",
        "内推",
        "https://example.com/jd/123",
        "2026-05-07",
        "投递",
        "5/14 未回则发邮件 follow up",
        "2026-05-14",
        "40-55k",
        "张三",
        "内推码 ABC123",
    ]
    ws.append(sample_row)

    # 下拉验证：来源（C 列）
    dv_source = DataValidation(
        type="list",
        formula1=f'"{",".join(SOURCE_OPTIONS)}"',
        allow_blank=True,
    )
    dv_source.add(f"C2:C200")
    ws.add_data_validation(dv_source)

    # 下拉验证：当前阶段（F 列）
    dv_stage = DataValidation(
        type="list",
        formula1=f'"{",".join(STAGE_OPTIONS)}"',
        allow_blank=True,
    )
    dv_stage.add(f"F2:F200")
    ws.add_data_validation(dv_stage)

    # 条件格式：F 列 = Offer 绿色，Reject 灰色，沉默 黄色
    ws.conditional_formatting.add(
        "F2:F200",
        CellIsRule(operator="equal", formula=['"Offer"'],
                   fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        "F2:F200",
        CellIsRule(operator="equal", formula=['"Reject"'],
                   fill=PatternFill("solid", fgColor="D9D9D9")),
    )
    ws.conditional_formatting.add(
        "F2:F200",
        CellIsRule(operator="equal", formula=['"沉默"'],
                   fill=PatternFill("solid", fgColor="FFEB9C")),
    )

    # 冻结首行
    ws.freeze_panes = "A2"

    # 第二个 sheet：使用说明
    legend = wb.create_sheet("使用说明")
    legend["A1"] = "投递追踪表使用说明"
    legend["A1"].font = Font(bold=True, size=14)
    legend_lines = [
        "",
        "1. 每投递一个岗位，新增一行。",
        "2. 「来源」「当前阶段」是下拉框，直接选。",
        "3. 「下一步动作」务必写具体可执行的事 + 时间点。",
        "4. 状态超过 7 天没动静时，把「当前阶段」改成『沉默』，提醒自己跟进或放弃。",
        "5. 拿到 Offer 后，行会变绿；被拒后会变灰，方便区分。",
        "6. 建议每周固定一个时间（比如周日晚）回顾一遍这张表。",
        "",
        "如果你想看『有多少在面试中 / 投递总数 / 沉默率』等统计，",
        "可以让 Claude 帮你做一个看板（artifact）。",
    ]
    for line in legend_lines:
        legend.append([line])

    wb.save(path)
    print(f"✓ Excel 追踪表已生成：{path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="初始化求职投递追踪表")
    parser.add_argument(
        "--format", choices=["xlsx", "md"], default="xlsx", help="输出格式"
    )
    parser.add_argument(
        "--output", required=True, help="输出文件路径（含文件名）"
    )
    args = parser.parse_args()

    out = Path(args.output).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)

    if args.format == "xlsx":
        write_xlsx(out)
    else:
        write_md(out)


if __name__ == "__main__":
    main()
